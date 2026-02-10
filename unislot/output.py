"""Output generation: Export schedule and clash report to Excel."""

from pathlib import Path
from typing import Optional
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from unislot.models import ClashReport, ClashStatus, Schedule, Day

# ============== Style Definitions ==============

# Color palette - Professional, clean colors
COLORS = {
    "primary": "0F172A",  # Dark slate - main headers
    "primary_light": "1E3A8A",  # Blue - secondary headers  
    "accent": "3B82F6",  # Bright blue - highlights
    "secondary": "059669",  # Emerald - success/good
    "success": "10B981",  # Green - positive indicators
    "success_bg": "ECFDF5",  # Light green background
    "warning": "F59E0B",  # Amber - warnings
    "warning_bg": "FFFBEB",  # Light amber background
    "danger": "DC2626",  # Red - errors/clashes
    "danger_bg": "FEF2F2",  # Light red background
    "header_bg": "1E293B",  # Dark header
    "subheader_bg": "334155",  # Medium header
    "row_alt": "F8FAFC",  # Alternating row - very light
    "row_hover": "F1F5F9",  # Slightly darker
    "border": "E2E8F0",  # Light border
    "border_dark": "94A3B8",  # Darker border for headers
    "text_primary": "0F172A",  # Main text
    "text_secondary": "64748B",  # Secondary text
    "white": "FFFFFF",
}

# Day colors - Subtle, professional tints
DAY_COLORS = {
    "Monday": "EFF6FF",  # Very light blue
    "Tuesday": "F0FDF4",  # Very light green
    "Wednesday": "FFFBEB",  # Very light yellow
    "Thursday": "FDF2F8",  # Very light pink
    "Friday": "EEF2FF",  # Very light indigo
    "Saturday": "FEF3C7",  # Very light amber (for math courses)
}


def _get_styles() -> dict:
    """Create reusable styles for Excel formatting."""
    thin_border = Border(left=Side(style='thin', color=COLORS["border"]),
                         right=Side(style='thin', color=COLORS["border"]),
                         top=Side(style='thin', color=COLORS["border"]),
                         bottom=Side(style='thin', color=COLORS["border"]))

    medium_border = Border(left=Side(style='medium',
                                     color=COLORS["border_dark"]),
                           right=Side(style='medium',
                                      color=COLORS["border_dark"]),
                           top=Side(style='medium',
                                    color=COLORS["border_dark"]),
                           bottom=Side(style='medium',
                                       color=COLORS["border_dark"]))

    return {
        "title": {
            "font":
            Font(bold=True, size=18, color=COLORS["white"]),
            "fill":
            PatternFill(start_color=COLORS["primary"],
                        end_color=COLORS["primary"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='center', vertical='center'),
            "border":
            medium_border,
        },
        "subtitle": {
            "font":
            Font(bold=True, size=13, color=COLORS["white"]),
            "fill":
            PatternFill(start_color=COLORS["subheader_bg"],
                        end_color=COLORS["subheader_bg"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='center', vertical='center'),
            "border":
            thin_border,
        },
        "section_header": {
            "font":
            Font(bold=True, size=12, color=COLORS["white"]),
            "fill":
            PatternFill(start_color=COLORS["primary_light"],
                        end_color=COLORS["primary_light"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='left', vertical='center', indent=1),
            "border":
            thin_border,
        },
        "header": {
            "font":
            Font(bold=True, size=10, color=COLORS["white"]),
            "fill":
            PatternFill(start_color=COLORS["header_bg"],
                        end_color=COLORS["header_bg"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='center', vertical='center', wrap_text=True),
            "border":
            thin_border,
        },
        "cell": {
            "font":
            Font(size=10, color=COLORS["text_primary"]),
            "alignment":
            Alignment(horizontal='left',
                      vertical='center',
                      wrap_text=True,
                      indent=1),
            "border":
            thin_border,
        },
        "cell_center": {
            "font": Font(size=10, color=COLORS["text_primary"]),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": thin_border,
        },
        "cell_number": {
            "font": Font(size=10, color=COLORS["text_primary"]),
            "alignment": Alignment(horizontal='right', vertical='center'),
            "border": thin_border,
        },
        "success": {
            "font":
            Font(size=10, bold=True, color=COLORS["secondary"]),
            "fill":
            PatternFill(start_color=COLORS["success_bg"],
                        end_color=COLORS["success_bg"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='center', vertical='center'),
            "border":
            thin_border,
        },
        "danger": {
            "font":
            Font(size=10, bold=True, color=COLORS["danger"]),
            "fill":
            PatternFill(start_color=COLORS["danger_bg"],
                        end_color=COLORS["danger_bg"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='center', vertical='center'),
            "border":
            thin_border,
        },
        "highlight_green": {
            "font":
            Font(size=10, color=COLORS["secondary"]),
            "fill":
            PatternFill(start_color=COLORS["success_bg"],
                        end_color=COLORS["success_bg"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='left', vertical='center', indent=1),
            "border":
            Border(left=Side(style='medium', color=COLORS["success"]),
                   right=Side(style='thin', color=COLORS["border"]),
                   top=Side(style='thin', color=COLORS["border"]),
                   bottom=Side(style='thin', color=COLORS["border"])),
        },
        "total_row": {
            "font":
            Font(bold=True, size=10, color=COLORS["text_primary"]),
            "fill":
            PatternFill(start_color=COLORS["row_hover"],
                        end_color=COLORS["row_hover"],
                        fill_type="solid"),
            "alignment":
            Alignment(horizontal='right', vertical='center'),
            "border":
            Border(left=Side(style='thin', color=COLORS["border"]),
                   right=Side(style='thin', color=COLORS["border"]),
                   top=Side(style='medium', color=COLORS["border_dark"]),
                   bottom=Side(style='medium', color=COLORS["border_dark"])),
        },
        "metric_label": {
            "font": Font(size=10, color=COLORS["text_secondary"]),
            "alignment": Alignment(horizontal='left',
                                   vertical='center',
                                   indent=1),
        },
        "metric_value": {
            "font": Font(bold=True, size=12, color=COLORS["primary_light"]),
            "alignment": Alignment(horizontal='left', vertical='center'),
        },
        "border": thin_border,
    }


def _apply_style(cell, style_dict: dict):
    """Apply style dictionary to a cell."""
    for key, value in style_dict.items():
        if key != "border":
            setattr(cell, key, value)
    if "border" in style_dict:
        cell.border = style_dict["border"]


def _set_column_width(ws, col_idx: int, width: int):
    """Set column width with minimum constraint."""
    ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 8)


def _freeze_panes(ws, row: int, col: int = 1):
    """Freeze panes for easier scrolling."""
    ws.freeze_panes = ws.cell(row=row, column=col)


def export_schedule_xlsx(
    schedule: Schedule,
    output_path: Path | str,
    department_name: str = "COMPUTING TECHNOLOGIES",
) -> Path:
    """
    Export schedule to Excel file in SRM format with comprehensive categorization.
    
    Creates multiple organized sheets:
    - Schedule: Main timetable in SRM format
    - By Day: Day-wise grouped view with subtotals
    - By Program: Courses categorized by program/branch
    - Course Catalog: All courses with complete details
    - Summary: Comprehensive statistics and analysis
    
    Args:
        schedule: Schedule object with entries
        output_path: Path for output file
        department_name: Department name for header
        
    Returns:
        Path to created file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _get_styles()

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # Always exists for new workbook
    ws.title = "Schedule"

    # ========== MAIN SCHEDULE SHEET ==========

    # Header rows (SRM format)
    headers = [
        "SRM INSTITUTE OF SCIENCE AND TECHNOLOGY",
        "COLLEGE OF ENGINEERING AND TECHNOLOGY",
        "COMPENSATORY COURSES - Even 2025-26",
        "TIME TABLE (UG & PG) _ R2021",
        f"NAME OF THE DEPARTMENT : {department_name}",
    ]

    # Column headers
    column_headers = [
        "S.NO", "BRANCH", "COURSE CODE", "COURSE TITLE",
        "Total No. of Students", "DAY", "TIMING", "VENUE", "FACULTY NAME",
        "Faculty ID No", "FACULTY MOBILE NO", "FACULTY Email"
    ]

    num_cols = len(column_headers)

    # Write header rows with enhanced styling
    for row_idx, header_text in enumerate(headers, start=1):
        ws.merge_cells(start_row=row_idx,
                       start_column=1,
                       end_row=row_idx,
                       end_column=num_cols)
        cell = ws.cell(row=row_idx, column=1, value=header_text)
        if row_idx <= 2:
            _apply_style(cell, styles["title"])
            ws.row_dimensions[row_idx].height = 30
        else:
            _apply_style(cell, styles["subtitle"])
            ws.row_dimensions[row_idx].height = 25

    # Write column headers
    header_row = len(headers) + 2  # Add a blank row
    ws.row_dimensions[header_row].height = 35

    for col_idx, header in enumerate(column_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        _apply_style(cell, styles["header"])

    # Sort entries by day for better organization
    day_order = {
        "Monday": 0,
        "Tuesday": 1,
        "Wednesday": 2,
        "Thursday": 3,
        "Friday": 4,
        "Saturday": 5,
    }
    sorted_entries = sorted(schedule.entries,
                            key=lambda e:
                            (day_order.get(e.day.value, 6), e.course_code))

    # Write data rows with alternating colors and day-based highlighting
    data_start_row = header_row + 1
    current_day = None

    for idx, entry in enumerate(sorted_entries, start=1):
        row = data_start_row + idx - 1
        ws.row_dimensions[row].height = 22

        # Get day color for row
        day_color = DAY_COLORS.get(entry.day.value, COLORS["row_alt"])
        row_fill = PatternFill(start_color=day_color,
                               end_color=day_color,
                               fill_type="solid")

        # Row data
        row_data = [
            idx,  # S.NO
            entry.programs,  # BRANCH
            entry.course_code,
            entry.course_title,
            entry.enrollment_count,
            entry.day.value,
            "05.00 PM to 07.00 PM",
            "",  # VENUE
            entry.faculty or "",  # FACULTY NAME
            "",  # Faculty ID
            "",  # MOBILE
            "",  # Email
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            _apply_style(
                cell, styles["cell_center"]
                if col_idx in [1, 5, 6] else styles["cell"])
            cell.fill = row_fill

    # Adjust column widths
    column_widths = [6, 35, 14, 50, 12, 12, 20, 10, 25, 14, 16, 28]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ========== DAY-WISE SCHEDULE SHEET ==========
    ws_days = wb.create_sheet("By Day")

    # Group entries by day
    day_groups = defaultdict(list)
    for entry in sorted_entries:
        day_groups[entry.day.value].append(entry)

    row = 1
    # Sheet title
    ws_days.merge_cells(start_row=row,
                        start_column=1,
                        end_row=row,
                        end_column=7)
    cell = ws_days.cell(row=row, column=1, value="SCHEDULE BY DAY")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_days.row_dimensions[row].height = 35
    row += 2

    for day in [
            "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"
    ]:
        entries = day_groups.get(day, [])
        if not entries:
            continue

        # Day header with count and total enrollment
        total_enrollment = sum(e.enrollment_count for e in entries)
        day_color = DAY_COLORS.get(day, COLORS["row_alt"])
        ws_days.merge_cells(start_row=row,
                            start_column=1,
                            end_row=row,
                            end_column=7)
        cell = ws_days.cell(
            row=row,
            column=1,
            value=
            f"{day.upper()} — {len(entries)} sections | {total_enrollment} students"
        )
        cell.font = Font(bold=True, size=13, color=COLORS["white"])
        cell.fill = PatternFill(start_color=COLORS["primary"],
                                end_color=COLORS["primary"],
                                fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws_days.row_dimensions[row].height = 28
        row += 1

        # Column headers for this day
        day_headers = [
            "#", "Course Code", "Course Title", "Section", "Enrollment",
            "Programs", "Time"
        ]
        for col_idx, header in enumerate(day_headers, start=1):
            cell = ws_days.cell(row=row, column=col_idx, value=header)
            _apply_style(cell, styles["header"])
        ws_days.row_dimensions[row].height = 24
        row += 1

        # Entries for this day
        for idx, entry in enumerate(entries, start=1):
            row_fill = PatternFill(start_color=day_color,
                                   end_color=day_color,
                                   fill_type="solid")
            day_row_data = [
                idx,
                entry.course_code,
                entry.course_title,
                entry.section_number,
                entry.enrollment_count,
                entry.programs[:40] +
                "..." if len(entry.programs) > 40 else entry.programs,
                entry.time,
            ]
            for col_idx, value in enumerate(day_row_data, start=1):
                cell = ws_days.cell(row=row, column=col_idx, value=value)
                _apply_style(
                    cell, styles["cell_center"]
                    if col_idx in [1, 4, 5] else styles["cell"])
                cell.fill = row_fill
            ws_days.row_dimensions[row].height = 20
            row += 1

        # Subtotal row
        ws_days.merge_cells(start_row=row,
                            start_column=1,
                            end_row=row,
                            end_column=4)
        cell = ws_days.cell(row=row,
                            column=1,
                            value=f"Subtotal: {len(entries)} sections")
        cell.font = Font(bold=True, size=10, color=COLORS["primary"])
        cell.alignment = Alignment(horizontal='right')
        ws_days.cell(row=row, column=5,
                     value=total_enrollment).font = Font(bold=True)
        ws_days.row_dimensions[row].height = 22
        row += 2

    # Adjust column widths for day sheet
    day_col_widths = [5, 14, 40, 10, 12, 35, 18]
    for col_idx, width in enumerate(day_col_widths, start=1):
        ws_days.column_dimensions[get_column_letter(col_idx)].width = width

    # ========== BY PROGRAM SHEET ==========
    ws_programs = wb.create_sheet("By Program")

    # Group by program
    program_groups = defaultdict(list)
    for entry in sorted_entries:
        # Parse programs from the comma-separated string
        programs = [p.strip() for p in entry.programs.split(",") if p.strip()]
        for program in programs:
            program_groups[program].append(entry)

    row = 1
    # Sheet title
    ws_programs.merge_cells(start_row=row,
                            start_column=1,
                            end_row=row,
                            end_column=6)
    cell = ws_programs.cell(row=row, column=1, value="SCHEDULE BY PROGRAM")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_programs.row_dimensions[row].height = 35
    row += 2

    for program in sorted(program_groups.keys()):
        entries = program_groups[program]

        # Program header
        ws_programs.merge_cells(start_row=row,
                                start_column=1,
                                end_row=row,
                                end_column=6)
        cell = ws_programs.cell(row=row,
                                column=1,
                                value=f"{program} — {len(entries)} courses")
        cell.font = Font(bold=True, size=12, color=COLORS["white"])
        cell.fill = PatternFill(start_color=COLORS["secondary"],
                                end_color=COLORS["secondary"],
                                fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws_programs.row_dimensions[row].height = 26
        row += 1

        # Headers
        prog_headers = [
            "#", "Course Code", "Course Title", "Day", "Time", "Enrollment"
        ]
        for col_idx, header in enumerate(prog_headers, start=1):
            cell = ws_programs.cell(row=row, column=col_idx, value=header)
            _apply_style(cell, styles["header"])
        ws_programs.row_dimensions[row].height = 22
        row += 1

        # Courses for this program
        for idx, entry in enumerate(sorted(entries,
                                           key=lambda e: e.course_code),
                                    start=1):
            day_color = DAY_COLORS.get(entry.day.value, COLORS["row_alt"])
            row_fill = PatternFill(start_color=day_color,
                                   end_color=day_color,
                                   fill_type="solid")

            row_data = [
                idx, entry.course_code, entry.course_title, entry.day.value,
                entry.time, entry.enrollment_count
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_programs.cell(row=row, column=col_idx, value=value)
                _apply_style(
                    cell, styles["cell_center"]
                    if col_idx in [1, 6] else styles["cell"])
                cell.fill = row_fill
            ws_programs.row_dimensions[row].height = 20
            row += 1
        row += 1

    # Adjust widths
    prog_widths = [5, 14, 45, 12, 18, 12]
    for col_idx, width in enumerate(prog_widths, start=1):
        ws_programs.column_dimensions[get_column_letter(col_idx)].width = width

    # ========== COURSE CATALOG SHEET ==========
    ws_catalog = wb.create_sheet("Course Catalog")

    row = 1
    # Sheet title
    ws_catalog.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=8)
    cell = ws_catalog.cell(row=row, column=1, value="COMPLETE COURSE CATALOG")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_catalog.row_dimensions[row].height = 35
    row += 2

    # Headers
    catalog_headers = [
        "S.No", "Course Code", "Course Title", "Total Sections",
        "Total Enrollment", "Scheduled Days", "Programs/Branches", "Faculty"
    ]
    for col_idx, header in enumerate(catalog_headers, start=1):
        cell = ws_catalog.cell(row=row, column=col_idx, value=header)
        _apply_style(cell, styles["header"])
    ws_catalog.row_dimensions[row].height = 28
    row += 1

    # Group by course code
    course_groups = defaultdict(list)
    for entry in schedule.entries:
        course_groups[entry.course_code].append(entry)

    day_order = [
        "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday",
        "Sunday"
    ]
    for idx, (course_code, entries) in enumerate(sorted(course_groups.items()),
                                                 start=1):
        total_enrollment = sum(e.enrollment_count for e in entries)
        days = sorted(set(e.day.value for e in entries),
                      key=lambda d: day_order.index(d)
                      if d in day_order else 99)
        programs = sorted(
            set(p.strip() for e in entries for p in e.programs.split(",")
                if p.strip()))
        faculty = sorted(set(e.faculty for e in entries if e.faculty))

        row_data = [
            idx,
            course_code,
            entries[0].course_title,
            len(entries),
            total_enrollment,
            ", ".join(days),
            ", ".join(programs[:3]) + ("..." if len(programs) > 3 else ""),
            ", ".join(faculty) if faculty else "—",
        ]

        row_fill = PatternFill(
            start_color=COLORS["row_alt"] if idx % 2 == 0 else COLORS["white"],
            end_color=COLORS["row_alt"] if idx % 2 == 0 else COLORS["white"],
            fill_type="solid")

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_catalog.cell(row=row, column=col_idx, value=value)
            _apply_style(
                cell, styles["cell_center"]
                if col_idx in [1, 4, 5] else styles["cell"])
            cell.fill = row_fill
        ws_catalog.row_dimensions[row].height = 22
        row += 1

    # Total row
    row += 1
    ws_catalog.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=3)
    ws_catalog.cell(row=row, column=1, value="TOTAL").font = Font(bold=True,
                                                                  size=11)
    ws_catalog.cell(row=row, column=4,
                    value=len(schedule.entries)).font = Font(bold=True)
    ws_catalog.cell(row=row,
                    column=5,
                    value=sum(
                        e.enrollment_count
                        for e in schedule.entries)).font = Font(bold=True)

    # Adjust widths
    catalog_widths = [6, 14, 40, 12, 14, 25, 35, 25]
    for col_idx, width in enumerate(catalog_widths, start=1):
        ws_catalog.column_dimensions[get_column_letter(col_idx)].width = width

    # ========== SUMMARY SHEET ==========
    ws_summary = wb.create_sheet("Summary")

    # Title
    ws_summary.merge_cells(start_row=1,
                           start_column=1,
                           end_row=1,
                           end_column=4)
    cell = ws_summary.cell(row=1, column=1, value="SCHEDULE SUMMARY")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 35

    row = 3

    # ===== OVERVIEW SECTION =====
    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    cell = ws_summary.cell(row=row, column=1, value="OVERVIEW")
    cell.font = Font(bold=True, size=13, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["secondary"],
                            end_color=COLORS["secondary"],
                            fill_type="solid")
    ws_summary.row_dimensions[row].height = 28
    row += 1

    overview_metrics = [
        ("Total Sections Scheduled", schedule.total_sections, "sections"),
        ("Total Student Enrollments",
         sum(e.enrollment_count for e in schedule.entries), "students"),
        ("Unique Courses", len(course_groups), "courses"),
        ("Unique Programs", len(program_groups), "programs"),
        ("Days Utilized", len([d for d in day_groups
                               if day_groups[d]]), "days"),
    ]

    for metric_name, value, unit in overview_metrics:
        ws_summary.cell(row=row, column=1,
                        value=metric_name).font = Font(size=11)
        value_cell = ws_summary.cell(row=row, column=2, value=value)
        value_cell.font = Font(bold=True,
                               size=12,
                               color=COLORS["primary_light"])
        ws_summary.cell(row=row, column=3,
                        value=unit).font = Font(size=10, color="666666")
        row += 1
    row += 1

    # ===== SOLVER DETAILS =====
    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    cell = ws_summary.cell(row=row, column=1, value="SOLVER DETAILS")
    cell.font = Font(bold=True, size=13, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["accent"],
                            end_color=COLORS["accent"],
                            fill_type="solid")
    ws_summary.row_dimensions[row].height = 28
    row += 1

    solver_metrics = [
        ("Algorithm Used", schedule.solver_used.upper()),
        ("Computation Time", f"{schedule.solver_time_seconds:.2f} seconds"),
        ("Students with Clashes", schedule.total_clashes),
        ("Clash Rate",
         f"{(schedule.total_clashes / sum(1 for _ in schedule.entries) * 100):.1f}%"
         if schedule.entries else "0%"),
    ]

    for metric_name, value in solver_metrics:
        ws_summary.cell(row=row, column=1,
                        value=metric_name).font = Font(size=11)
        value_cell = ws_summary.cell(row=row, column=2, value=value)
        color = COLORS[
            "danger"] if "Clashes" in metric_name and schedule.total_clashes > 0 else COLORS[
                "success"] if "Clashes" in metric_name else COLORS["accent"]
        value_cell.font = Font(bold=True, size=11, color=color)
        row += 1
    row += 1

    # ===== DAY DISTRIBUTION =====
    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    cell = ws_summary.cell(row=row, column=1, value="DAY DISTRIBUTION")
    cell.font = Font(bold=True, size=13, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary_light"],
                            end_color=COLORS["primary_light"],
                            fill_type="solid")
    ws_summary.row_dimensions[row].height = 28
    row += 1

    # Header row for day distribution
    for col_idx, header in enumerate(["Day", "Sections", "Students", "Visual"],
                                     start=1):
        ws_summary.cell(row=row, column=col_idx,
                        value=header).font = Font(bold=True, size=10)
    row += 1

    for day in [
            "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"
    ]:
        entries = day_groups.get(day, [])
        count = len(entries)
        students = sum(e.enrollment_count for e in entries)

        day_color = DAY_COLORS.get(day, COLORS["row_alt"])
        row_fill = PatternFill(start_color=day_color,
                               end_color=day_color,
                               fill_type="solid")

        ws_summary.cell(row=row, column=1, value=day).fill = row_fill
        ws_summary.cell(row=row, column=2, value=count).font = Font(bold=True)
        ws_summary.cell(row=row, column=3, value=students)

        # Visual bar
        bar_length = min(count, 25)
        bar = "█" * bar_length + ("+" if count > 25 else "")
        ws_summary.cell(row=row, column=4,
                        value=bar).font = Font(size=9, color=COLORS["primary"])
        row += 1

    # Total row
    ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row, column=2,
                    value=len(schedule.entries)).font = Font(bold=True)
    ws_summary.cell(row=row,
                    column=3,
                    value=sum(
                        e.enrollment_count
                        for e in schedule.entries)).font = Font(bold=True)
    row += 2

    # ===== TOP COURSES =====
    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    cell = ws_summary.cell(row=row,
                           column=1,
                           value="TOP COURSES BY ENROLLMENT")
    cell.font = Font(bold=True, size=13, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["warning"],
                            end_color=COLORS["warning"],
                            fill_type="solid")
    ws_summary.row_dimensions[row].height = 28
    row += 1

    # Get top 10 courses by enrollment
    course_enrollments = [(code, sum(e.enrollment_count for e in entries),
                           entries[0].course_title)
                          for code, entries in course_groups.items()]
    top_courses = sorted(course_enrollments, key=lambda x: x[1],
                         reverse=True)[:10]

    for col_idx, header in enumerate(
        ["Rank", "Course Code", "Title", "Enrollment"], start=1):
        ws_summary.cell(row=row, column=col_idx,
                        value=header).font = Font(bold=True, size=10)
    row += 1

    for rank, (code, enrollment, title) in enumerate(top_courses, start=1):
        ws_summary.cell(row=row, column=1, value=rank)
        ws_summary.cell(row=row, column=2, value=code).font = Font(bold=True)
        ws_summary.cell(row=row,
                        column=3,
                        value=title[:35] + "..." if len(title) > 35 else title)
        ws_summary.cell(row=row, column=4, value=enrollment).font = Font(
            bold=True, color=COLORS["primary_light"])
        row += 1

    # Adjust column widths
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 40
    ws_summary.column_dimensions["D"].width = 30

    wb.save(output_path)
    return output_path


def export_schedule_simple_xlsx(schedule: Schedule,
                                output_path: Path | str) -> Path:
    """
    Export schedule to simple Excel file with modern styling.
    
    Creates a clean, well-formatted schedule without SRM-specific headers.
    
    Args:
        schedule: Schedule object with entries
        output_path: Path for output file
        
    Returns:
        Path to created file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _get_styles()
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # Always exists for new workbook
    ws.title = "Schedule"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    cell = ws.cell(row=1, column=1, value="COURSE SCHEDULE")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Headers
    headers = [
        "#", "Course Code", "Course Title", "Section", "Day", "Time",
        "Enrollment"
    ]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        _apply_style(cell, styles["header"])
    ws.row_dimensions[row].height = 28

    # Sort by day
    day_order = {
        "Monday": 0,
        "Tuesday": 1,
        "Wednesday": 2,
        "Thursday": 3,
        "Friday": 4,
        "Saturday": 5,
    }
    sorted_entries = sorted(schedule.entries,
                            key=lambda e:
                            (day_order.get(e.day.value, 6), e.course_code))

    # Data rows
    row = 4
    for idx, entry in enumerate(sorted_entries, start=1):
        day_color = DAY_COLORS.get(entry.day.value, COLORS["row_alt"])
        row_fill = PatternFill(start_color=day_color,
                               end_color=day_color,
                               fill_type="solid")

        row_data = [
            idx,
            entry.course_code,
            entry.course_title,
            entry.section_number,
            entry.day.value,
            entry.time,
            entry.enrollment_count,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            _apply_style(
                cell, styles["cell_center"]
                if col_idx in [1, 4, 7] else styles["cell"])
            cell.fill = row_fill

        ws.row_dimensions[row].height = 22
        row += 1

    # Column widths
    col_widths = [5, 14, 45, 10, 12, 18, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")

    ws_summary.merge_cells(start_row=1,
                           start_column=1,
                           end_row=1,
                           end_column=2)
    cell = ws_summary.cell(row=1, column=1, value="SUMMARY")
    cell.font = Font(bold=True, size=14, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws_summary.row_dimensions[1].height = 30

    metrics = [
        ("Total Sections", schedule.total_sections),
        ("Solver Used", schedule.solver_used.upper()),
        ("Solver Time", f"{schedule.solver_time_seconds:.2f}s"),
        ("Total Clashes", schedule.total_clashes),
    ]

    row = 3
    for metric_name, value in metrics:
        ws_summary.cell(row=row, column=1, value=metric_name)
        ws_summary.cell(row=row, column=2, value=value).font = Font(bold=True)
        row += 1

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15

    wb.save(output_path)
    return output_path


def export_clash_report_xlsx(report: ClashReport,
                             output_path: Path | str) -> Path:
    """
    Export clash report to Excel file with comprehensive categorization.
    
    Creates organized sheets:
    - Summary: Comprehensive overview with detailed statistics
    - Clashes Only: Students with scheduling conflicts
    - By Program: Clashes categorized by program/branch
    - By Day: Clashes grouped by the day they occur
    - By Course: Analysis of which courses cause most clashes
    - Full Report: Complete student listing with status
    
    Args:
        report: ClashReport object
        output_path: Path for output file
        
    Returns:
        Path to created file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _get_styles()
    wb = Workbook()

    # Get students with clashes for analysis
    clashing_students = [
        r for r in report.reports if r.status == ClashStatus.RED
    ]

    # ========== SUMMARY SHEET (First) ==========
    ws_summary = wb.active
    assert ws_summary is not None  # Always exists for new workbook
    ws_summary.title = "Summary"

    # Title
    ws_summary.merge_cells(start_row=1,
                           start_column=1,
                           end_row=1,
                           end_column=4)
    cell = ws_summary.cell(row=1, column=1, value="CLASH REPORT")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 35

    row = 3

    # ===== STATUS INDICATOR =====
    status_text = "ALL CLEAR — No Scheduling Conflicts!" if report.students_with_clashes == 0 else f"ATTENTION REQUIRED — {report.students_with_clashes} Students Have Conflicts"
    status_color = COLORS[
        "success"] if report.students_with_clashes == 0 else COLORS["danger"]

    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    status_cell = ws_summary.cell(row=row, column=1, value=status_text)
    status_cell.font = Font(bold=True, size=14, color=status_color)
    status_cell.alignment = Alignment(horizontal='center')
    ws_summary.row_dimensions[row].height = 32
    row += 2

    # ===== KEY METRICS SECTION =====
    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    cell = ws_summary.cell(row=row, column=1, value="KEY METRICS")
    cell.font = Font(bold=True, size=13, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["secondary"],
                            end_color=COLORS["secondary"],
                            fill_type="solid")
    ws_summary.row_dimensions[row].height = 28
    row += 1

    metrics = [
        ("Total Students Analyzed", report.total_students, "students"),
        ("Students with Clashes", report.students_with_clashes, "students"),
        ("Clash-Free Students", report.clash_free_students, "students"),
        ("Clash Rate", f"{report.clash_percentage:.2f}%", ""),
        ("Success Rate", f"{100 - report.clash_percentage:.2f}%", ""),
    ]

    for metric_name, value, unit in metrics:
        ws_summary.cell(row=row, column=1,
                        value=metric_name).font = Font(size=11)
        value_cell = ws_summary.cell(row=row, column=2, value=value)

        if "Clashes" in metric_name and isinstance(value, int):
            color = COLORS["danger"] if value > 0 else COLORS["success"]
        elif "Success" in metric_name or "Free" in metric_name:
            color = COLORS["success"]
        else:
            color = COLORS["primary_light"]
        value_cell.font = Font(bold=True, size=12, color=color)
        ws_summary.cell(row=row, column=3,
                        value=unit).font = Font(size=10, color="666666")
        row += 1
    row += 1

    if clashing_students:
        # ===== CLASH ANALYSIS SECTION =====
        ws_summary.merge_cells(start_row=row,
                               start_column=1,
                               end_row=row,
                               end_column=4)
        cell = ws_summary.cell(row=row, column=1, value="CLASH ANALYSIS")
        cell.font = Font(bold=True, size=13, color=COLORS["white"])
        cell.fill = PatternFill(start_color=COLORS["danger"],
                                end_color=COLORS["danger"],
                                fill_type="solid")
        ws_summary.row_dimensions[row].height = 28
        row += 1

        # Analyze clashes by day
        day_clashes = defaultdict(int)
        for student in clashing_students:
            if student.clashing_day:
                day_clashes[student.clashing_day.value] += 1

        ws_summary.cell(row=row, column=1,
                        value="Clashes by Day:").font = Font(bold=True,
                                                             size=11)
        row += 1
        for day in [
                "Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                "Saturday"
        ]:
            count = day_clashes.get(day, 0)
            if count > 0:
                ws_summary.cell(row=row, column=1, value=f"  • {day}")
                ws_summary.cell(row=row, column=2, value=count).font = Font(
                    bold=True, color=COLORS["danger"])
                ws_summary.cell(row=row, column=3, value="students")
                row += 1
        row += 1

        # Analyze clashes by program
        program_clashes = defaultdict(int)
        for student in clashing_students:
            program_clashes[student.program] += 1

        ws_summary.cell(row=row, column=1,
                        value="Clashes by Program:").font = Font(bold=True,
                                                                 size=11)
        row += 1
        for program, count in sorted(program_clashes.items(),
                                     key=lambda x: x[1],
                                     reverse=True)[:5]:
            ws_summary.cell(row=row, column=1, value=f"  • {program[:30]}...")
            ws_summary.cell(row=row, column=2,
                            value=count).font = Font(bold=True,
                                                     color=COLORS["danger"])
            row += 1
        row += 1

        # Analyze which course pairs cause most clashes
        course_pair_counts = defaultdict(int)
        for student in clashing_students:
            for pair in student.clashing_courses:
                sorted_pair = tuple(sorted(pair))
                course_pair_counts[sorted_pair] += 1

        ws_summary.cell(row=row,
                        column=1,
                        value="Most Problematic Course Pairs:").font = Font(
                            bold=True, size=11)
        row += 1
        for (c1, c2), count in sorted(course_pair_counts.items(),
                                      key=lambda x: x[1],
                                      reverse=True)[:5]:
            ws_summary.cell(row=row, column=1, value=f"  • {c1} & {c2}")
            ws_summary.cell(row=row, column=2,
                            value=count).font = Font(bold=True,
                                                     color=COLORS["danger"])
            ws_summary.cell(row=row, column=3, value="students affected")
            row += 1

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 20
    ws_summary.column_dimensions["D"].width = 20

    # ========== CLASHES ONLY SHEET ==========
    ws_clashes = wb.create_sheet("Clashes Only")

    if clashing_students:
        # Title
        ws_clashes.merge_cells(start_row=1,
                               start_column=1,
                               end_row=1,
                               end_column=7)
        cell = ws_clashes.cell(
            row=1,
            column=1,
            value=
            f"STUDENTS WITH SCHEDULING CONFLICTS ({len(clashing_students)})")
        cell.font = Font(bold=True, size=14, color=COLORS["white"])
        cell.fill = PatternFill(start_color=COLORS["danger"],
                                end_color=COLORS["danger"],
                                fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_clashes.row_dimensions[1].height = 32

        # Headers
        row = 3
        headers = [
            "S.No", "Register No.", "Student Name", "Program",
            "Enrolled Courses", "Clashing Courses", "Clash Day"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_clashes.cell(row=row, column=col_idx, value=header)
            _apply_style(cell, styles["header"])
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       wrap_text=False)
        ws_clashes.row_dimensions[row].height = 28
        row += 1

        # Sort by program, then name
        sorted_clashers = sorted(clashing_students,
                                 key=lambda s: (s.program, s.student_name))

        # Data rows
        for idx, student in enumerate(sorted_clashers, start=1):
            clashing_text = "; ".join(f"{c1} & {c2}"
                                      for c1, c2 in student.clashing_courses)
            enrolled_text = ", ".join(student.enrolled_courses)

            row_data = [
                idx,
                student.register_number,
                student.student_name,
                student.program,
                enrolled_text,
                clashing_text,
                student.clashing_day.value if student.clashing_day else "",
            ]

            row_fill = PatternFill(start_color="FEE2E2",
                                   end_color="FEE2E2",
                                   fill_type="solid")
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_clashes.cell(row=row, column=col_idx, value=value)
                _apply_style(
                    cell,
                    styles["cell_center"] if col_idx == 1 else styles["cell"])
                cell.fill = row_fill
                # Prevent text wrapping causing misalignment
                if col_idx not in [1, 7]:  # Not S.No or Clash Day
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=False)

            ws_clashes.row_dimensions[row].height = 22
            row += 1

        # Adjust widths - wider columns to prevent cutoff
        clash_widths = [6, 18, 28, 50, 30, 32, 14]
        for col_idx, width in enumerate(clash_widths, start=1):
            ws_clashes.column_dimensions[get_column_letter(
                col_idx)].width = width
    else:
        ws_clashes.merge_cells(start_row=1,
                               start_column=1,
                               end_row=1,
                               end_column=3)
        cell = ws_clashes.cell(
            row=1,
            column=1,
            value="No clashes! All students have conflict-free schedules.")
        cell.font = Font(bold=True, size=14, color=COLORS["success"])
        cell.alignment = Alignment(horizontal='center')

    # ========== BY PROGRAM SHEET ==========
    ws_by_program = wb.create_sheet("By Program")

    row = 1
    ws_by_program.merge_cells(start_row=row,
                              start_column=1,
                              end_row=row,
                              end_column=6)
    cell = ws_by_program.cell(row=row, column=1, value="CLASHES BY PROGRAM")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_by_program.row_dimensions[row].height = 35
    row += 2

    if clashing_students:
        # Group by program
        program_groups = defaultdict(list)
        for student in clashing_students:
            program_groups[student.program].append(student)

        for program in sorted(program_groups.keys()):
            students = program_groups[program]

            # Program header
            ws_by_program.merge_cells(start_row=row,
                                      start_column=1,
                                      end_row=row,
                                      end_column=6)
            cell = ws_by_program.cell(
                row=row,
                column=1,
                value=f"{program} — {len(students)} students with clashes")
            cell.font = Font(bold=True, size=12, color=COLORS["white"])
            cell.fill = PatternFill(start_color=COLORS["secondary"],
                                    end_color=COLORS["secondary"],
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws_by_program.row_dimensions[row].height = 26
            row += 1

            # Headers
            prog_headers = [
                "#", "Register No.", "Name", "Clashing Courses", "Day"
            ]
            for col_idx, header in enumerate(prog_headers, start=1):
                cell = ws_by_program.cell(row=row,
                                          column=col_idx,
                                          value=header)
                _apply_style(cell, styles["header"])
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=False)
            ws_by_program.row_dimensions[row].height = 24
            row += 1

            for idx, student in enumerate(sorted(students,
                                                 key=lambda s: s.student_name),
                                          start=1):
                clash_text = "; ".join(f"{c1} & {c2}"
                                       for c1, c2 in student.clashing_courses)
                row_data = [
                    idx, student.register_number, student.student_name,
                    clash_text,
                    student.clashing_day.value if student.clashing_day else ""
                ]

                row_fill = PatternFill(start_color="FEE2E2",
                                       end_color="FEE2E2",
                                       fill_type="solid")
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_by_program.cell(row=row,
                                              column=col_idx,
                                              value=value)
                    _apply_style(
                        cell, styles["cell_center"]
                        if col_idx == 1 else styles["cell"])
                    cell.fill = row_fill
                    # Prevent text wrapping
                    if col_idx not in [1, 5]:
                        cell.alignment = Alignment(horizontal='left',
                                                   vertical='center',
                                                   wrap_text=False)
                ws_by_program.row_dimensions[row].height = 22
                row += 1
            row += 1

        # Widths - wider for course codes
        prog_widths = [5, 18, 28, 35, 14]
        for col_idx, width in enumerate(prog_widths, start=1):
            ws_by_program.column_dimensions[get_column_letter(
                col_idx)].width = width
    else:
        ws_by_program.cell(
            row=row, column=1,
            value="No clashes to display.").font = Font(italic=True)

    # ========== BY DAY SHEET ==========
    ws_by_day = wb.create_sheet("By Day")

    row = 1
    ws_by_day.merge_cells(start_row=row,
                          start_column=1,
                          end_row=row,
                          end_column=5)
    cell = ws_by_day.cell(row=row, column=1, value="CLASHES BY DAY")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_by_day.row_dimensions[row].height = 35
    row += 2

    if clashing_students:
        day_groups = defaultdict(list)
        for student in clashing_students:
            if student.clashing_day:
                day_groups[student.clashing_day.value].append(student)

        for day in [
                "Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                "Saturday"
        ]:
            students = day_groups.get(day, [])
            if not students:
                continue

            day_color = DAY_COLORS.get(day, COLORS["row_alt"])

            # Day header
            ws_by_day.merge_cells(start_row=row,
                                  start_column=1,
                                  end_row=row,
                                  end_column=5)
            cell = ws_by_day.cell(
                row=row,
                column=1,
                value=f"{day.upper()} — {len(students)} clashes")
            cell.font = Font(bold=True, size=13, color=COLORS["white"])
            cell.fill = PatternFill(start_color=COLORS["primary"],
                                    end_color=COLORS["primary"],
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws_by_day.row_dimensions[row].height = 28
            row += 1

            # Headers
            day_headers = [
                "#", "Register No.", "Name", "Program", "Clashing Courses"
            ]
            for col_idx, header in enumerate(day_headers, start=1):
                cell = ws_by_day.cell(row=row, column=col_idx, value=header)
                _apply_style(cell, styles["header"])
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=False)
            ws_by_day.row_dimensions[row].height = 24
            row += 1

            for idx, student in enumerate(students, start=1):
                clash_text = "; ".join(f"{c1} & {c2}"
                                       for c1, c2 in student.clashing_courses)
                row_data = [
                    idx, student.register_number, student.student_name,
                    student.program, clash_text
                ]

                row_fill = PatternFill(start_color=day_color,
                                       end_color=day_color,
                                       fill_type="solid")
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_by_day.cell(row=row, column=col_idx, value=value)
                    _apply_style(
                        cell, styles["cell_center"]
                        if col_idx == 1 else styles["cell"])
                    cell.fill = row_fill
                    # Prevent text wrapping
                    if col_idx != 1:
                        cell.alignment = Alignment(horizontal='left',
                                                   vertical='center',
                                                   wrap_text=False)
                ws_by_day.row_dimensions[row].height = 22
                row += 1
            row += 1

        # Widths - wider for program and courses
        day_widths = [5, 18, 28, 50, 35]
        for col_idx, width in enumerate(day_widths, start=1):
            ws_by_day.column_dimensions[get_column_letter(
                col_idx)].width = width
    else:
        ws_by_day.cell(row=row, column=1,
                       value="No clashes to display.").font = Font(italic=True)

    # ========== BY COURSE SHEET ==========
    ws_by_course = wb.create_sheet("By Course")

    row = 1
    ws_by_course.merge_cells(start_row=row,
                             start_column=1,
                             end_row=row,
                             end_column=4)
    cell = ws_by_course.cell(row=row,
                             column=1,
                             value="CLASH ANALYSIS BY COURSE PAIR")
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_by_course.row_dimensions[row].height = 35
    row += 2

    if clashing_students:
        # Analyze course pairs
        course_pair_students = defaultdict(list)
        for student in clashing_students:
            for pair in student.clashing_courses:
                sorted_pair = tuple(sorted(pair))
                course_pair_students[sorted_pair].append(student)

        # Headers
        course_headers = [
            "Rank", "Course Pair", "Students Affected", "Programs Affected"
        ]
        for col_idx, header in enumerate(course_headers, start=1):
            cell = ws_by_course.cell(row=row, column=col_idx, value=header)
            _apply_style(cell, styles["header"])
        ws_by_course.row_dimensions[row].height = 26
        row += 1

        for rank, ((c1, c2),
                   students) in enumerate(sorted(course_pair_students.items(),
                                                 key=lambda x: len(x[1]),
                                                 reverse=True),
                                          start=1):
            programs = sorted(set(s.program for s in students))

            row_data = [
                rank,
                f"{c1} & {c2}",
                len(students),
                ", ".join(programs[:3]) + ("..." if len(programs) > 3 else ""),
            ]

            # Color intensity based on severity
            intensity = min(255, 200 + len(students) * 5)
            row_fill = PatternFill(start_color="FEE2E2",
                                   end_color="FEE2E2",
                                   fill_type="solid")

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_by_course.cell(row=row, column=col_idx, value=value)
                _apply_style(
                    cell, styles["cell_center"]
                    if col_idx in [1, 3] else styles["cell"])
                cell.fill = row_fill
            ws_by_course.row_dimensions[row].height = 22
            row += 1

        # Widths
        course_widths = [6, 35, 18, 45]
        for col_idx, width in enumerate(course_widths, start=1):
            ws_by_course.column_dimensions[get_column_letter(
                col_idx)].width = width
    else:
        ws_by_course.cell(
            row=row, column=1,
            value="No clashes to analyze.").font = Font(italic=True)

    # ========== FULL REPORT SHEET ==========
    ws_full = wb.create_sheet("Full Report")

    # Title
    ws_full.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    cell = ws_full.cell(row=1,
                        column=1,
                        value="COMPLETE STUDENT STATUS REPORT")
    cell.font = Font(bold=True, size=14, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_full.row_dimensions[1].height = 32

    # Headers
    row = 3
    headers = [
        "S.No", "Register No.", "Student Name", "Program", "Enrolled Courses",
        "Status", "Clash Details"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_full.cell(row=row, column=col_idx, value=header)
        _apply_style(cell, styles["header"])
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=False)
    ws_full.row_dimensions[row].height = 28
    row += 1

    # Sort: clashers first, then by program, then alphabetically
    sorted_reports = sorted(
        report.reports,
        key=lambda r: (r.status != ClashStatus.RED, r.program, r.student_name))

    for idx, student in enumerate(sorted_reports, start=1):
        clashing_text = "; ".join(
            f"{c1} & {c2}" for c1, c2 in
            student.clashing_courses) if student.clashing_courses else "—"
        status_text = "CLASH" if student.status == ClashStatus.RED else "OK"

        row_data = [
            idx,
            student.register_number,
            student.student_name,
            student.program,
            ", ".join(student.enrolled_courses),
            status_text,
            clashing_text,
        ]

        # Row color based on status
        if student.status == ClashStatus.RED:
            row_fill = PatternFill(start_color="FEE2E2",
                                   end_color="FEE2E2",
                                   fill_type="solid")
        elif idx % 2 == 0:
            row_fill = PatternFill(start_color=COLORS["row_alt"],
                                   end_color=COLORS["row_alt"],
                                   fill_type="solid")
        else:
            row_fill = PatternFill(start_color=COLORS["white"],
                                   end_color=COLORS["white"],
                                   fill_type="solid")

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_full.cell(row=row, column=col_idx, value=value)
            _apply_style(
                cell,
                styles["cell_center"] if col_idx in [1, 6] else styles["cell"])
            cell.fill = row_fill
            # Ensure text doesn't wrap unexpectedly - keep single line
            if col_idx not in [1, 6]:  # Not S.No or Status
                cell.alignment = Alignment(horizontal='left',
                                           vertical='center',
                                           wrap_text=False)

        ws_full.row_dimensions[row].height = 22
        row += 1

    # Adjust widths - wider columns to prevent cutoff
    full_widths = [6, 18, 28, 50, 30, 10, 32]
    for col_idx, width in enumerate(full_widths, start=1):
        ws_full.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path


def export_fixed_schedule_xlsx(
    presorted_schedule,  # PresortedSchedule from parser
    applied_fixes: list[dict],  # List of {course_code, new_day} dicts
    output_path: Path | str,
) -> Path:
    """
    Export a fixed schedule to Excel with changed rows highlighted in green.
    
    Creates an Excel file similar to the uploaded faculty format, but with
    the applied fixes incorporated and the changed rows highlighted.
    
    Args:
        presorted_schedule: Original PresortedSchedule from parser
        applied_fixes: List of dicts with 'course_code' and 'new_day' keys
        output_path: Path for output file
        
    Returns:
        Path to created file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    styles = _get_styles()

    # Build a map of course -> new_day for quick lookup
    fix_map: dict[str, str] = {}
    for fix in applied_fixes:
        fix_map[fix["course_code"]] = fix["new_day"]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fixed Schedule"

    # Header row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    title_cell = ws.cell(row=row, column=1, value="FIXED COURSE SCHEDULE")
    title_cell.font = Font(bold=True, size=16, color=COLORS["white"])
    title_cell.fill = PatternFill(start_color=COLORS["primary"],
                                  end_color=COLORS["primary"],
                                  fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 35

    # Subtitle showing fixes count
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    subtitle_cell = ws.cell(
        row=row,
        column=1,
        value=
        f"Generated with {len(applied_fixes)} course reassignments highlighted in green"
    )
    subtitle_cell.font = Font(size=11,
                              italic=True,
                              color=COLORS["text_secondary"])
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25

    row += 2

    # Column headers
    headers = [
        "#", "Register No", "Student Name", "Program", "Course Code",
        "Course Title", "Day"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        _apply_style(cell, styles["header"])
    ws.row_dimensions[row].height = 28

    row += 1
    data_start_row = row

    # Get student data and course-day mapping
    students = presorted_schedule.students
    original_course_day_map = presorted_schedule.course_day_map.copy()

    # Apply the fixes to get the new course-day mapping
    fixed_course_day_map = original_course_day_map.copy()
    for course_code, new_day in fix_map.items():
        fixed_course_day_map[course_code] = new_day

    # Green highlight style
    green_fill = PatternFill(start_color=COLORS["success_bg"],
                             end_color=COLORS["success_bg"],
                             fill_type="solid")
    green_border = Border(left=Side(style='medium', color=COLORS["success"]),
                          right=Side(style='thin', color=COLORS["border"]),
                          top=Side(style='thin', color=COLORS["border"]),
                          bottom=Side(style='thin', color=COLORS["border"]))

    idx = 0
    if students:
        # Sort students by register number for consistent output
        sorted_students = sorted(students.values(),
                                 key=lambda s: s.register_number)

        for student in sorted_students:
            for course_code in student.enrolled_courses:
                idx += 1

                # Check if this course was changed
                is_changed = course_code in fix_map
                day = fixed_course_day_map.get(course_code, "Unknown")

                # Get course title from entries if available
                course_title = ""
                for entry in presorted_schedule.entries:
                    if entry.course_code == course_code:
                        course_title = entry.course_title
                        break

                row_data = [
                    idx,
                    student.register_number,
                    student.name,
                    student.program,
                    course_code,
                    course_title,
                    day,
                ]

                # Alternating row colors (unless changed)
                if idx % 2 == 0:
                    row_fill = PatternFill(start_color=COLORS["row_alt"],
                                           end_color=COLORS["row_alt"],
                                           fill_type="solid")
                else:
                    row_fill = PatternFill(start_color=COLORS["white"],
                                           end_color=COLORS["white"],
                                           fill_type="solid")

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    _apply_style(
                        cell, styles["cell_center"]
                        if col_idx in [1, 7] else styles["cell"])

                    if is_changed:
                        cell.fill = green_fill
                        if col_idx == 1:  # First column gets green left border
                            cell.border = green_border
                        if col_idx == 7:  # Day column - make it bold
                            cell.font = Font(bold=True,
                                             size=10,
                                             color=COLORS["secondary"])
                    else:
                        cell.fill = row_fill

                ws.row_dimensions[row].height = 20
                row += 1

    # Column widths
    widths = [6, 15, 25, 28, 15, 35, 12]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = f"A{data_start_row}"

    # Add a summary sheet
    ws_summary = wb.create_sheet(title="Changes Summary")

    row = 1
    ws_summary.merge_cells(start_row=row,
                           start_column=1,
                           end_row=row,
                           end_column=4)
    cell = ws_summary.cell(row=row, column=1, value="APPLIED CHANGES")
    cell.font = Font(bold=True, size=14, color=COLORS["white"])
    cell.fill = PatternFill(start_color=COLORS["primary"],
                            end_color=COLORS["primary"],
                            fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[row].height = 30

    row += 2

    # Headers
    summary_headers = ["#", "Course Code", "Original Day", "New Day"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=row, column=col_idx, value=header)
        _apply_style(cell, styles["header"])
    ws_summary.row_dimensions[row].height = 25

    row += 1

    # List applied fixes
    for idx, fix in enumerate(applied_fixes, start=1):
        course_code = fix["course_code"]
        original_day = original_course_day_map.get(course_code, "Unknown")
        new_day = fix["new_day"]

        row_data = [idx, course_code, original_day, new_day]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row, column=col_idx, value=value)
            _apply_style(
                cell,
                styles["cell_center"] if col_idx == 1 else styles["cell"])

            # Highlight the new day in green
            if col_idx == 4:
                cell.fill = green_fill
                cell.font = Font(bold=True, size=10, color=COLORS["secondary"])

        ws_summary.row_dimensions[row].height = 22
        row += 1

    # Summary widths
    for col_idx, width in enumerate([6, 15, 15, 15], start=1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    return output_path


def export_all(
    schedule: Schedule,
    clash_report: ClashReport,
    output_dir: Path | str,
) -> tuple[Path, Path]:
    """
    Export both schedule and clash report.
    
    Args:
        schedule: Schedule object
        clash_report: ClashReport object
        output_dir: Directory for output files
        
    Returns:
        Tuple of (schedule_path, clash_report_path)
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    schedule_path = export_schedule_xlsx(schedule,
                                         output_dir / "schedule.xlsx")
    clash_path = export_clash_report_xlsx(clash_report,
                                          output_dir / "clash_report.xlsx")

    return schedule_path, clash_path
